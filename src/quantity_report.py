from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from .models import TakeoffItem


class QuantityReporter:
    def __init__(self, output_dir: Path) -> None:
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_json(self, items: Iterable[TakeoffItem], path: Path) -> None:
        path.write_text(json.dumps([item.to_dict() for item in items], indent=2), encoding="utf-8")

    def write_excel(self, items: list[TakeoffItem], review_items: list, assumptions: list[str], path: Path) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            path.with_suffix(".json").write_text(json.dumps([item.to_dict() for item in items], indent=2), encoding="utf-8")
            return

        wb = Workbook()
        category = wb.active
        category.title = "Summary by Category"
        category.append(["Category", "Item", "Unit", "Included Quantity", "Review Quantity", "Average Confidence"])

        grouped: dict[tuple[str, str, str], list[TakeoffItem]] = defaultdict(list)
        for item in items:
            grouped[(item.category, item.item_name, item.unit)].append(item)

        for key, values in sorted(grouped.items()):
            included = sum(value.quantity or 0 for value in values if value.include_in_final)
            review_qty = sum(value.quantity or 0 for value in values if not value.include_in_final and value.quantity is not None)
            average_confidence = sum(value.confidence for value in values) / len(values)
            category.append([*key, round(included, 2), round(review_qty, 2), round(average_confidence, 2)])

        by_sheet = wb.create_sheet("Summary by Sheet")
        by_sheet.append(["Sheet", "Category", "Item", "Unit", "Included Quantity"])
        sheet_group: dict[tuple[str | None, str, str, str], float] = defaultdict(float)
        for item in items:
            if item.include_in_final and item.quantity is not None:
                sheet_group[(item.sheet_number, item.category, item.item_name, item.unit)] += item.quantity
        for key, quantity in sorted(sheet_group.items(), key=lambda row: tuple(str(value) for value in row[0])):
            by_sheet.append([*key, round(quantity, 2)])

        detail = wb.create_sheet("Detailed Items")
        detail.append(["Category", "Item", "Sheet", "Type", "Quantity", "Unit", "Source", "Confidence", "Included", "Review", "Notes"])
        for item in items:
            detail.append(
                [
                    item.category,
                    item.item_name,
                    item.sheet_number,
                    item.measurement_type,
                    item.quantity,
                    item.unit,
                    item.source_method,
                    item.confidence,
                    item.include_in_final,
                    item.review_required,
                    item.notes,
                ]
            )

        review = wb.create_sheet("Low Confidence Review")
        review.append(["Kind", "Sheet", "Category", "Item/Reason", "Quantity", "Unit", "Confidence", "Notes"])
        for entry in review_items:
            if isinstance(entry, TakeoffItem):
                review.append([entry.source_method, entry.sheet_number, entry.category, entry.item_name, entry.quantity, entry.unit, entry.confidence, entry.notes])
            else:
                review.append([entry.get("type"), entry.get("sheet_number"), None, entry.get("reason"), None, None, None, entry.get("sheet_id")])

        assumptions_sheet = wb.create_sheet("Assumptions")
        assumptions_sheet.append(["Assumption"])
        for assumption in assumptions:
            assumptions_sheet.append([assumption])

        wb.save(path)

    def write_pdf(self, project: str, items: list[TakeoffItem], review_items: list, assumptions: list[str], path: Path) -> None:
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            path.with_suffix(".txt").write_text(self._plain_report(project, items, review_items, assumptions), encoding="utf-8")
            return

        pdf = canvas.Canvas(str(path), pagesize=letter)
        _, height = letter
        y = height - 48
        for line in self._plain_report(project, items, review_items, assumptions).splitlines():
            if y < 48:
                pdf.showPage()
                y = height - 48
            pdf.drawString(48, y, line[:115])
            y -= 14
        pdf.save()

    def _plain_report(self, project: str, items: list[TakeoffItem], review_items: list, assumptions: list[str]) -> str:
        lines = [
            f"Auto Takeoff Report: {project}",
            "",
            f"Detected items: {len(items)}",
            f"Review-needed entries: {len(review_items)}",
            "",
            "Included Quantity Summary",
        ]
        grouped: dict[tuple[str, str, str], float] = defaultdict(float)
        for item in items:
            if item.include_in_final and item.quantity is not None:
                grouped[(item.category, item.item_name, item.unit)] += item.quantity
        for (category, name, unit), quantity in sorted(grouped.items()):
            lines.append(f"- {category} | {name}: {quantity:.2f} {unit}")
        lines.extend(["", "Warnings and Assumptions"])
        lines.extend(f"- {assumption}" for assumption in assumptions[:100])
        return "\n".join(lines)
